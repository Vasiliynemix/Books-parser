__doc__ = '\nФункции работы с файлами и папками.\n'

import inspect
import os
import re
import sys
from contextlib import suppress

import openpyxl

MAX_DIR_NAME_LENGTH = 255


def adjust_column_widths(excel_filepath: str):
    """Автоматически подгоняет ширину столбцов в файле Excel."""
    workbook = openpyxl.load_workbook(excel_filepath)
    worksheet = workbook.active

    # Создаем стиль для границ
    border_style = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

    # Создаем стиль выравнивания (по левому краю для ячеек, кроме заголовков)
    left_alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

    # Создаем стиль для выравнивания заголовков по центру
    center_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for column_number, column_cells in enumerate(worksheet.columns, 1):
        max_column_width = 0

        # Пробегаем по всем ячейкам в столбце и ищем максимальную длину строки
        for cell in column_cells:
            column_width = len(str(cell.value))
            if column_width > max_column_width:
                max_column_width = column_width

        # Добавляем немного ширины столбцу (для отступов)
        max_column_width += 2

        # Для столбца "Цена" добавляем немного ширины
        if column_number == 23:
            max_column_width += 6

        # Для столбца "Описание" ограничиваем ширину
        if column_number == 22:
            max_column_width = min(max_column_width, 20)  # Ограничиваем ширину 20 символами

        # Устанавливаем ширину столбца
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_column_width

        # Применяем выравнивание и границы для всех ячеек в столбце
        for row_number, cell in enumerate(column_cells):
            cell.border = border_style

            # Если это первая строка (заголовки), выравниваем по центру
            if row_number == 0:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    workbook.save(excel_filepath)


def get_data_dirpath() -> str:
    """Формирует путь к каталогу данных для работы программы."""
    return os.path.join(get_script_dirpath(), '.data')


def get_excel_filepath(shop: str, publisher: str) -> str:
    """Формирует путь к Exсel файлу с информацией о книгах."""
    return os.path.join(get_shop_publisher_dirpath(shop, publisher), f'content information {normalize_dir_name(publisher)}.xlsx')


def get_icon_filepath() -> str:
    """Формирует путь к файлу к иконке программы."""
    return os.path.join(get_script_dirpath(), 'logo.ico')


def get_images_dirpath(shop: str, publisher: str) -> str:
    """Формирует путь к каталогу с картинками обложек книг."""
    return os.path.join(get_shop_publisher_dirpath(shop, publisher), 'Images')


def get_missing_images_dirpath(shop: str, publisher: str) -> str:
    """
    Формирует путь к каталогу с теми картинками обложек книг, которые означают,
    что реальная картинка обложки отсутствует.
    """
    return os.path.join(get_shop_publisher_dirpath(shop, publisher), 'missing images')


def cut_shop_to_site(shop: str) -> str:
    """Обрезает название магазина до имени сайта."""
    return shop[:shop.find('.')]


def get_publishers_filepath(shop: str) -> str:
    """Формирует путь к файлу для сохранения издательств."""
    publishers_filename = f"{cut_shop_to_site(shop)}_publishers.txt"
    return os.path.join(get_data_dirpath(), publishers_filename)


def get_script_dirpath(follow_symlinks: bool = True) -> str:
    """Вычисляет путь к каталогу, из которого запущен данный скрипт."""
    if getattr(sys, 'frozen', False):
        path = os.path.abspath(sys.executable)
    else:
        path = inspect.getabsfile(get_script_dirpath)
    if follow_symlinks:
        path = os.path.realpath(path)
    return os.path.dirname(path)


def get_shop_dirpath(shop: str) -> str:
    """Формирует путь к каталогу конкретного магазина."""
    return os.path.join(get_shops_dirpath(), shop)


def get_xml_dirpath(shop: str) -> str:
    """Формирует путь к .xml файлу конкретного магазина."""
    return os.path.join(get_shop_dirpath(shop), f'{cut_shop_to_site(shop)}.xml')


def get_shop_publisher_dirpath(shop: str, publisher: str) -> str:
    """Формирует путь к каталогу магазина и издательства."""
    shop_publisher_filename = normalize_dir_name(f"{shop}-{publisher.strip()}")
    return os.path.join(get_shop_dirpath(shop), shop_publisher_filename)


def get_shops_dirpath() -> str:
    """Формирует путь к общему каталогу магазинов."""
    return os.path.join(get_script_dirpath(), 'content of shops')


def make_invisible_dir(dirpath: str) -> None:
    """Создаёт невидимый каталог."""
    os.makedirs(dirpath, exist_ok=True)
    os.system(f'attrib +h "{dirpath}"')


def normalize_dir_name(dir_name: str) -> str:
    """
    Приводит имя папки к допустимому в Windows имени папки:
       — удаляет запрещённые символы,
       — удаляет точку в конце строки,
       — усекает строку до максимально допустимой длины 255 символов.
    """
    dir_name = re.sub('[<>:"/\\\\|?*]', '', dir_name)
    if dir_name.endswith('.'):
        dir_name = dir_name[:-1]
    if len(dir_name) > MAX_DIR_NAME_LENGTH:
        dir_name = dir_name[:MAX_DIR_NAME_LENGTH]
    return dir_name


def is_file_exists(filepath):
    return os.path.exists(filepath)


def prepare_excel_file(excel_filepath: str) -> None:
    """
    Делает Excel файл пустым или создаёт его.
    Записывает в первой строке файла заголовки столбцов
    полужирным шрифтом и обводит ячейки с ними рамками.
    """

    if os.path.exists(excel_filepath):
        # Delete file
        os.remove(excel_filepath)

        # workbook = openpyxl.load_workbook(excel_filepath)
        # worksheet = workbook.worksheets[0]
        # for row in worksheet.iter_rows():
        #     for cell in row:
        #         cell.value = None

    # else:
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.append([
        'ISBN', 'Название', 'Название языка', 'Серия',
        'Название издательства или производителя', 'Авторы',
        'Категория товара', 'Возрастная категория', 'Переплёт',
        'Страна производитель', 'Год издания', 'Вид товара',
        'Количество страниц', 'Размеры', 'Длина', 'Ширина',
        'Высота', 'Класс', 'Вес', 'Цвет', 'Тип бумаги', 'Описание', 'Цена', 'ID'
    ])
    first_row = worksheet[1]
    alignment = openpyxl.styles.Alignment(horizontal='center',
                                          vertical='center')
    for cell in first_row:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                             right=openpyxl.styles.Side(style='thin'),
                                             top=openpyxl.styles.Side(style='thin'),
                                             bottom=openpyxl.styles.Side(style='thin'))
        cell.alignment = alignment

    workbook.save(excel_filepath)


def prepare_shop_dirs(shop):
    shops_dirpath = get_shops_dirpath()
    os.makedirs(shops_dirpath, exist_ok=True)
    shop_dirpath = get_shop_dirpath(shop)
    os.makedirs(shop_dirpath, exist_ok=True)
    # create_logger(shop_dirpath + f'/logs_{shop}.txt')
    # create_logger(shop_dirpath + f'/logs.txt')


def prepare_output_dirs_and_files(shop: str, publisher: str) -> None:
    """Создаёт нужные каталоги и файлы для загрузки книг."""
    prepare_shop_dirs(shop)
    shop_publisher_dirpath = get_shop_publisher_dirpath(shop, publisher)
    # if rmtree:  # remove directory and all its contents if set to True
    #     shutil.rmtree(shop_publisher_dirpath, ignore_errors=True)
    os.makedirs(shop_publisher_dirpath, exist_ok=True)
    images_dirpath = get_images_dirpath(shop, publisher)
    os.makedirs(images_dirpath, exist_ok=True)
    missing_images_dirpath = get_missing_images_dirpath(shop, publisher)
    # shutil.rmtree(missing_images_dirpath, ignore_errors=True)
    os.makedirs(missing_images_dirpath, exist_ok=True)
    excel_filepath = get_excel_filepath(shop, publisher)
    prepare_excel_file(excel_filepath)


def read_publishers(shop: str):
    """Читает список издательств магазина из файла."""
    publishers = []

    publishers_filepath = get_publishers_filepath(shop)
    with suppress(FileNotFoundError):
        with open(publishers_filepath, 'r', encoding='utf-8') as publishers_file:
            publishers = publishers_file.read().split('\n')

    return publishers


def get_books_rows_from_excel(excel_filepath, row_id):
    """Получает информацию об ISBN книг в файле Excel."""
    workbook = openpyxl.load_workbook(excel_filepath)
    worksheet = workbook.active

    isbn = []
    for row in worksheet.iter_rows(2, worksheet.max_row):
        isbn.append(str(row[row_id].value))
    return isbn


def try_to_change_type(obj, new_type):
    try:
        new_obj = new_type(obj)
    except (ValueError, TypeError):
        new_obj = obj
    return new_obj


def write_books_to_excel(excel_filepath: str, books: list):
    """Сохраняет информацию о книгах в файле Excel."""
    workbook = openpyxl.load_workbook(excel_filepath)
    worksheet = workbook.active
    start_row = worksheet.max_row + 1
    for book in books:
        # print("description " + book.get("description", " "))
        isbn = try_to_change_type(book.get('isbn', ' '), int)
        year = try_to_change_type(book.get('year', ' '), int)
        price = try_to_change_type(book.get('price', ' '), float)
        age_category = try_to_change_type(book.get('age_category', ' '), int)
        pages = try_to_change_type(book.get('pages', ' '), int)
        weight = try_to_change_type(book.get('weight', ' '), int)
        b_id = try_to_change_type(book.get('ID', ' '), int)
        grade = try_to_change_type(book.get('grade', ' '), int)

        worksheet.append([
            isbn, book.get('name', ' '), book.get('language', ' '),
            book.get('series', ' '), book.get('publisher', ' '),
            book.get('authors', ' '), book.get('category', ' '),
            age_category, book.get('cover', ' '),
            book.get('country', ' '), year, book.get('type', ' '),
            pages, book.get('dimensions', ' '),
            book.get('length', ' '), book.get('width', ' '),
            book.get('height', ' '), grade,
            weight, book.get('color', ' '),
            book.get('paper', ' '),
            book.get('description', ' '),
            price, b_id
        ])

    left_alignment = openpyxl.styles.Alignment(horizontal='left', vertical='bottom')
    center_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='bottom')
    for row in worksheet.iter_rows(min_row=start_row, max_row=(worksheet.max_row), min_col=1, max_col=24):
        for cell_number, cell in enumerate(row, 1):
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                 right=openpyxl.styles.Side(style='thin'),
                                                 top=openpyxl.styles.Side(style='thin'),
                                                 bottom=openpyxl.styles.Side(style='thin'))
            if cell_number in (1, 11):
                with suppress(ValueError, TypeError):
                    if str(cell.value):
                        if int(cell.value):
                            cell.number_format = '0'
                cell.alignment = left_alignment
            else:
                if cell_number == 23:
                    with suppress(ValueError):
                        if str(cell.value):
                            if float(cell.value):
                                cell.number_format = '#,##0.00 _?'
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment

    workbook.save(excel_filepath)


def write_publishers(shop: str, publishers: list) -> None:
    """Записывает список издательств в файл издательств магазина."""
    make_invisible_dir(get_data_dirpath())
    publishers_filepath = get_publishers_filepath(shop)
    with open(publishers_filepath, 'w', encoding='utf-8') as publishers_file:
        publishers_file.write('\n'.join(publishers))


def save_xml_to_file(shop: str, request):
    """Сохраняет xml-файл со всей БД в файл."""
    xml_filename = get_xml_dirpath(shop)

    # Save orig temp file
    with open(xml_filename, "w", encoding="utf-8") as file:
        file.write(request.text)


def get_correct_xml_content(shop: str) -> str:
    """Возвращает содержимое xml-файла, исправляя кодировку."""
    with open(get_xml_dirpath(shop), "r", encoding="utf-8") as f:
        orig_content = f.read()

    content = orig_content.replace('encoding="windows-1251"', 'encoding="utf-8"')
    return content


def sanitize_filename(name: str) -> str:
    """Удаляет или заменяет недопустимые символы в имени файла."""
    return re.sub(r'[\/:*?"<>|]', '', name)
